# -*- coding: utf-8 -*-
"""
Created on Tue Dec 17 14:57:00 2024

@author: Hsin.YH.Yang
"""

import os
import numpy as np
import pandas as pd
from itertools import product
from openpyxl import load_workbook, Workbook
from datetime import datetime

# %%
data_path = [r"D:\BenQ_Project\01_UR_lab\2024_11 Shanghai CS Major\4. Statistics\1. SmallFlick\All_SmallFlick_data_1218.xlsx",
             ]
file_sheet = ["arm_motion", "FingerIncludeAngle", "FingerAngle",
              "MedFreq", "iMVC_cal", "iMVCslope"]

subject = ["S01", "S02", "S03", "S04",
           "S05", "S06", "S07", "S08",
           "S09", "S10", "S11", "S12"]

mouse = ["A", "C", "EC2", "HS"]

axis_dir = ["elbow_x", "elbow_y", "elbow_z",
            "wrist_x", "wrist_y", "wrist_z"]

direct = ["+", "-"]

# 使用 itertools.product 生成所有組合
combinations = list(product(subject, mouse))
# 获取当前日期和时间
now = datetime.now()
# 将日期转换为指定格式
formatted_date = now.strftime("%m-%d-%H%M")


# %%



for file in data_path:
    save_name, extension = os.path.splitext(file)
    save_name = save_name + "_ed.xlsx"
    
    if os.path.exists(save_name):
        workbook = load_workbook(save_name)
        print(f"檔案 {save_name} 已存在，成功打開。")
    else:
        workbook = Workbook()
        sheet = workbook.active
        workbook.save(save_name)
        
        workbook = load_workbook(save_name)
    
    for ind_sheet in file_sheet:
        data = pd.read_excel(file,
                             sheet_name = ind_sheet)
        # 生成一個資料儲存的地方
        pos_data = pd.DataFrame(np.zeros([len(combinations), len(data.columns)]),
                                 columns = data.columns)
        
        for idx in range(len(combinations)):
            if "method" in data.columns:
                filtered_df = data[(data['subject'] == combinations[idx][0]) & \
                                   (data['mouse'] == combinations[idx][1]) & \
                                    (data['method'] == "mean")]
            else:
                filtered_df = data[(data['subject'] == combinations[idx][0]) & \
                                   (data['mouse'] == combinations[idx][1])]
            if filtered_df.index.any():
                for column in filtered_df.columns:
                    # print(column)
                    if type(filtered_df.loc[filtered_df.index[0], column]) != str and \
                        type(filtered_df.loc[filtered_df.index[0], column]) != bool:
                        pos_data.loc[idx, column] = np.mean(filtered_df.loc[filtered_df.index, column])
                    else:
                        pos_data.loc[idx, column] = filtered_df.loc[filtered_df.index[0], column]
        # 將 pos_data 寫入新分頁
        with pd.ExcelWriter(save_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            pos_data.to_excel(writer, sheet_name=ind_sheet, index=False)
 
# %%
subject = ["S01", "S02", "S03", "S04",
           "S05", "S06", "S07", "S08",
           "S09", "S10", "S11", "S12"]

mouse = ["A", "C", "EC2", "HS"]

# axis_dir = ["elbow_x", "elbow_y", "elbow_z",
#             "wrist_x", "wrist_y", "wrist_z"]

axis_dir = ["elbow", "wrist"]

direct = ["+", "-"]

method_dir = ["vel", "acc"]

# 使用 itertools.product 生成所有組合
combinations = list(product(subject, mouse, axis_dir, method_dir))

# %%

data_path = r"D:\BenQ_Project\01_UR_lab\2024_11 Shanghai CS Major\4. Statistics\3. SpiderShot\All_Spider_vicon_motion+emg_short.xlsx"


save_name, extension = os.path.splitext(data_path)
save_name = save_name + "_eded.xlsx"
    
if os.path.exists(save_name):
    workbook = load_workbook(save_name)
    print(f"檔案 {save_name} 已存在，成功打開。")
else:
    workbook = Workbook()
    sheet = workbook.active
    workbook.save(save_name)
    
    workbook = load_workbook(save_name)
    

data = pd.read_excel(data_path,
                     sheet_name = "工作表2")
# 生成一個資料儲存的地方
pos_data = pd.DataFrame(np.zeros([len(combinations), len(data.columns)]),
                            columns = data.columns)
        
for idx in range(len(combinations)):
    filtered_df = data[(data['subject'] == combinations[idx][0]) & \
                       (data['mouse'] == combinations[idx][1]) & \
                       (data['位置'] == combinations[idx][2]) & \
                    (data['method'] == combinations[idx][3])]
    if filtered_df.index.any():
        for column in filtered_df.columns:
            # print(column)
            if type(filtered_df.loc[filtered_df.index[0], column]) != str and \
                type(filtered_df.loc[filtered_df.index[0], column]) != bool:
                pos_data.loc[idx, column] = np.mean(filtered_df.loc[filtered_df.index, column])
            else:
                pos_data.loc[idx, column] = filtered_df.loc[filtered_df.index[0], column]
# 將 pos_data 寫入新分頁
with pd.ExcelWriter(save_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    pos_data.to_excel(writer, sheet_name="Sheet", index=False)
   
        
# %%

import pandas as pd
import pingouin as pg

from collections import defaultdict



data_path = [r"D:\BenQ_Project\01_UR_lab\2024_11 Shanghai CS Major\4. Statistics\1. SmallFlick\All_SmallFlick_data_1218_ed.xlsx",
             ]
data_path = r"D:\BenQ_Project\01_UR_lab\2024_11 Shanghai CS Major\4. Statistics\1. SmallFlick\All_SmallFlick_data_1218_ed.xlsx"
file_sheet = ["arm_motion", "FingerIncludeAngle", "FingerAngle",
              "MedFreq", "iMVC_cal", "iMVCslope"]


mouse = ["A", "C", "EC2", "HS"]

X = defaultdict(list)  # 預設為 list，所以不用手動初始化 X["A"] = []
# 建立一個儲存 ANOVA 結果的字典
anova_store = {}
posthoc_store = {}

for ind_sheet in file_sheet:
    data_table = pd.read_excel(data_path,
                               sheet_name=ind_sheet)
    # 建立資料儲存位置
    anova_store[ind_sheet] = {}  # 為每個 condition 建立字典
    posthoc_store[ind_sheet] = {}
    anova_store_pd = pd.DataFrame()
    posthoc_store_pd = pd.DataFrame()
    for column in data_table.columns[5:]:
        if data_table[column].isna().sum() < 1:
            # 執行 Repeated Measures ANOVA
            print(data_table[column])
            anova_results = pg.rm_anova(dv=column, within='mouse', subject='subject', data=data_table, detailed=True)
            anova_results["condition"] = column
            # 將每個條件的 anova results 合併
            anova_store_pd = pd.concat([anova_store_pd, anova_results])
            # print(anova_results)
            # 事後檢定（Post-hoc tests）
            posthoc = pg.pairwise_ttests(dv=column, within='mouse', subject='subject', data=data_table, padjust='bonf')
            posthoc["condition"] = column
            # 將每個條件的 posthoc results 合併
            posthoc_store_pd = pd.concat([posthoc_store_pd, posthoc])
        # print(posthoc)
        
    anova_store[ind_sheet] = anova_store_pd
    posthoc_store[ind_sheet] = posthoc_store_pd

# 指定要寫入的 Excel 檔案名稱
file_name = "anova_results.xlsx"

# 使用 `ExcelWriter` 寫入多個分頁
with pd.ExcelWriter(file_name, engine="xlsxwriter") as writer:
    for sheet_name, df in anova_store.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)  # `index=False` 避免寫入 DataFrame 索引



        


    

                    



















